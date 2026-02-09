from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator
import unicodedata

from openpyxl import load_workbook


EPS_NAME_MAP = {
    "SANITAS": "EPS SANITAS",
}


def _clean_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value)


def _strip_accents(value: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch)
    )


def normalize_eps_name(name: str | None) -> str | None:
    if not name:
        return None
    base = name.replace(".xlsx", "").strip()
    base = " ".join(base.split())
    return EPS_NAME_MAP.get(base, base)


def _find_sheet(wb, contains: str, endswith: str | None = None) -> str:
    contains_norm = _strip_accents(contains).lower()
    ends_norm = _strip_accents(endswith).lower() if endswith else None
    for name in wb.sheetnames:
        norm = _strip_accents(name).lower()
        if contains_norm in norm and (ends_norm is None or norm.endswith(ends_norm)):
            return name
    raise KeyError(f"No sheet matching {contains} {endswith or ''}".strip())


def _find_header_row(ws, expected_first_col: str, search_rows: int = 8) -> int:
    expected = _strip_accents(expected_first_col).upper()
    for r in range(1, search_rows + 1):
        cell = ws.cell(row=r, column=1).value
        if isinstance(cell, str):
            if _strip_accents(cell).upper() == expected:
                return r
    return 1


def load_eps_financials(path: str | Path, sheet: str | None = None) -> list[dict[str, object]]:
    """Load EPS financials table in wide format."""
    wb = load_workbook(Path(path), data_only=True)
    sheet_name = sheet or _find_sheet(wb, "eps")
    ws = wb[sheet_name]

    header_row = _find_header_row(ws, "EPS")
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    while headers and headers[-1] is None:
        headers.pop()

    records: list[dict[str, object]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if all(v is None for v in row):
            continue
        record = dict(zip(headers, row))
        for k, v in list(record.items()):
            if isinstance(v, str):
                record[k] = v.strip()
        if "EPS" in record:
            record["EPS_NORMALIZADA"] = normalize_eps_name(record.get("EPS"))
        records.append(record)
    return records


def financials_long(records: Iterable[dict[str, object]]) -> list[dict[str, object]]:
    """Normalize EPS financials to long format: eps, account, year, value."""
    rows: list[dict[str, object]] = []
    for record in records:
        eps = record.get("EPS")
        account = record.get("CUENTA")
        for key, value in record.items():
            if key in ("EPS", "CUENTA", "EPS_NORMALIZADA"):
                continue
            if not str(key).isdigit():
                continue
            if value is None:
                continue
            rows.append(
                {
                    "eps": eps,
                    "account": account,
                    "year": int(key),
                    "value": float(value),
                }
            )
    return rows


def load_caracterizacion02(
    path: str | Path,
    department: str | None = None,
    regimen: str | None = None,
) -> list[dict[str, object]]:
    """Load the Caracterizacion 02 sheet with affiliates by age group."""
    wb = load_workbook(Path(path), data_only=True)
    sheet_name = _find_sheet(wb, "caracteriz", "02")
    ws = wb[sheet_name]

    header_row = _find_header_row(ws, "DEPARTAMENTO")
    records: list[dict[str, object]] = []

    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
        if all(v is None for v in row_vals):
            continue
        dept, eps, reg, age, fem, male, norep, total = row_vals
        dept = _clean_str(dept)
        eps = _clean_str(eps)
        reg = _clean_str(reg)
        age = _clean_str(age)

        if department and dept != department:
            continue
        if regimen and reg != regimen:
            continue

        records.append(
            {
                "department": dept,
                "eps": eps,
                "eps_normalized": normalize_eps_name(eps),
                "regimen": reg,
                "age_group": age,
                "female": int(fem or 0),
                "male": int(male or 0),
                "no_report": int(norep or 0),
                "total": int(total or 0),
            }
        )
    return records


def age_group_share(
    records: Iterable[dict[str, object]],
    mode: str = "within_age",
) -> list[dict[str, object]]:
    """
    Compute percentages from afiliates by age group.

    mode:
      - within_age: percent of each EPS within an age group (EPS share by age).
      - within_eps: percent of each age group inside an EPS (age mix per EPS).
    """
    if mode not in {"within_age", "within_eps"}:
        raise ValueError("mode must be 'within_age' or 'within_eps'")

    totals_by_age: dict[tuple[str | None, str | None, str], int] = {}
    totals_by_eps: dict[tuple[str | None, str | None, str], int] = {}

    for r in records:
        age_key = (r.get("department"), r.get("regimen"), r.get("age_group"))
        eps_key = (r.get("department"), r.get("regimen"), r.get("eps"))
        totals_by_age[age_key] = totals_by_age.get(age_key, 0) + int(r.get("total") or 0)
        totals_by_eps[eps_key] = totals_by_eps.get(eps_key, 0) + int(r.get("total") or 0)

    rows: list[dict[str, object]] = []
    for r in records:
        age_key = (r.get("department"), r.get("regimen"), r.get("age_group"))
        eps_key = (r.get("department"), r.get("regimen"), r.get("eps"))
        denom = totals_by_age[age_key] if mode == "within_age" else totals_by_eps[eps_key]
        share = (r.get("total") or 0) / denom if denom else 0
        rows.append(
            {
                "department": r.get("department"),
                "regimen": r.get("regimen"),
                "eps": r.get("eps"),
                "age_group": r.get("age_group"),
                "total": r.get("total"),
                "share": share,
                "mode": mode,
            }
        )
    return rows
